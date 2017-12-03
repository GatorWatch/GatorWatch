from openpyxl import load_workbook
import datetime

# Show listing object to hold show information
class ShowListing:
    name = ""
    episode_name = ""
    episode = ""
    description = ""
    channel = ""
    date = ""
    time = ""

    def __init__(self, name, episode_name, episode, description, channel, date, time):
        self.name = name
        self.episode_name = episode_name
        self.episode = episode
        self.description = description
        self.channel = channel
        self.date = date
        self.time = time


# When the user wants to save a listing, save it to the calendar
def saveCalendar(listing):
    i = 2
    # Find an empty cell
    while ws.cell(row=i, column=1).value is not None:
        # There is already an event at that time - cannot save the given event
        if ws.cell(row=i, column=6).value == listing.date and ws.cell(row=i, column=7) == listing.time:
            return ws.cell(row=i,column=1).value
        i += 1

    ws.cell(row=i, column=1).value = listing.name
    ws.cell(row=i, column=2).value = listing.episode_name
    ws.cell(row=i, column=3).value = listing.episode
    ws.cell(row=i, column=4).value = listing.description
    ws.cell(row=i, column=5).value = listing.channel
    ws.cell(row=i, column=6).value = listing.date
    ws.cell(row=i, column=7).value = listing.time
    wb.save(filename="calendar_times.xlsx")
    return "True"   # Returning a string of True in case there is an issue with equality operators

# Return array of all the saved listings
def getCalendar():
    i = 2
    array = []

    while ws.cell(row=i, column=1).value is not None:
        # Extract values
        name = ws.cell(row=i, column=1).value
        episode_name = ws.cell(row=i, column=2).value
        episode_number = ws.cell(row=i, column=3).value
        description = ws.cell(row=i, column=4).value
        channel = ws.cell(row=i, column=5).value
        date = ws.cell(row=i, column=6).value
        time = ws.cell(row=i, column=7).value
        listing = ShowListing(name, episode_name, episode_number, description, channel, date, time)
        array.append(listing)

        i += 1

    return array

# Delete an event given the day and time
def deleteEvent(day, time):
    i = 0
    while ws.cell(row=i, column=1).value is not None:
        if ws.cell(row=i, column=6).value == day and ws.cell(row=i, column=7).value == time:
            break
        i += 1


    # Delete the listing
    ws.cell(row=i, column=1).value = None
    ws.cell(row=i, column=2).value = None
    ws.cell(row=i, column=3).value = None
    ws.cell(row=i, column=4).value = None
    ws.cell(row=i, column=5).value = None
    ws.cell(row=i, column=6).value = None
    ws.cell(row=i, column=7).value = None

    i += 1

    # Shift the values
    while ws.cell(row=i, column=1).value is not None:
        ws.cell(row=i-1, column=1).value = ws.cell(row=i, column=1).value
        ws.cell(row=i-1, column=2).value = ws.cell(row=i, column=2).value
        ws.cell(row=i-1, column=3).value = ws.cell(row=i, column=3).value
        ws.cell(row=i-1, column=4).value = ws.cell(row=i, column=4).value
        ws.cell(row=i-1, column=5).value = ws.cell(row=i, column=5).value
        ws.cell(row=i-1, column=6).value = ws.cell(row=i, column=6).value
        ws.cell(row=i-1, column=7).value = ws.cell(row=i, column=7).value
        i += 1

    # Delete the last listing since it is a duplicate
    i -= 1
    ws.cell(row=i, column=1).value = None
    ws.cell(row=i, column=2).value = None
    ws.cell(row=i, column=3).value = None
    ws.cell(row=i, column=4).value = None
    ws.cell(row=i, column=5).value = None
    ws.cell(row=i, column=6).value = None
    ws.cell(row=i, column=7).value = None

    wb.save(filename="calendar_times.xlsx")
    return True

# Clear the entire calendar of events
def clearCalendar():
    i = 2
    while ws.cell(row=i, column=1).value is not None:
        ws.cell(row=i, column=1).value = None
        ws.cell(row=i, column=2).value = None
        ws.cell(row=i, column=3).value = None
        ws.cell(row=i, column=4).value = None
        ws.cell(row=i, column=5).value = None
        ws.cell(row=i, column=6).value = None
        ws.cell(row=i, column=7).value = None
        wb.save("calendar_times.xlsx")
        i += 1

# Check to see if it is thirty minutes before an event - if there is one, return the event - else return false
def checkTime():
    months = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6, "Jul": 7,
              "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

    now = datetime.datetime.now()

    i = 2
    while ws.cell(row=i, column=1).value is not None:
        day = ws.cell(row=i, column=6).value
        time = ws.cell(row=i, column=7).value

        day = day.split()
        day[0] = months[day[0]]

        if day[0] != now.month and day[1] != now.day:
            i += 1
            continue

        time = time.split(":")
        hour = time[0]
        minutes = time[1]

        if minutes[2:] == "pm" and hour != 12:
            hour += 12

        elif minutes[2:] == "am" and hour == 12:
            hour = 0

        minutes = minutes[0:2]
        minutes = int(minutes)
        reminder_minutes = minutes - 30

        hour = int(hour)
        if reminder_minutes < 0:
            hour -= 1
            minutes = 60 + reminder_minutes

        if now.hour == hour and now.minute == minutes:
            name = ws.cell(row=i, column=1).value
            episode_name = ws.cell(row=i, column=2).value
            episode_number = ws.cell(row=i, column=3).value
            description = ws.cell(row=i, column=4).value
            channel = ws.cell(row=i, column=5).value
            date = ws.cell(row=i, column=6).value
            time = ws.cell(row=i, column=7).value
            listing = ShowListing(name, episode_name, episode_number, description, channel, date, time)
            return listing

        i += 1

    return False


wb = load_workbook("calendar_times.xlsx")
ws = wb["Sheet1"]